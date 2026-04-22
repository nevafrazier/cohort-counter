import pandas as pd
import re
import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FOLDER = os.path.dirname(os.path.abspath(__file__))


def get_latest_xlsx():
    files = glob.glob(os.path.join(FOLDER, '*.xlsx'))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def pct(n, total):
    return round(n / total * 100, 1) if total else 0


def run(filepath):
    xl = pd.ExcelFile(filepath)
    w_sheets = [s for s in xl.sheet_names if re.match(r'^W\d+$', s, re.IGNORECASE)]

    if not w_sheets:
        print("No sheets matching W# pattern (e.g. W20, W21) found.")
        sys.exit(1)

    results = []

    for sheet in w_sheets:
        df = pd.read_excel(filepath, sheet_name=sheet, header=0)

        if df.shape[1] < 9:
            print(f"[{sheet}] Skipped — no column I found.")
            continue

        col_i = df.iloc[:, 8].astype(str).str.strip()
        total = len(col_i)

        col_i_lower = col_i.str.lower()
        delivered        = int(col_i_lower.str.contains('delivered', na=False).sum())
        awaiting_pickup  = int(col_i_lower.str.contains('awaitingpickup|awaiting pickup|awaiting_pickup', na=False).sum())
        out_for_delivery = int(col_i_lower.str.contains('outfordelivery|out for delivery|out_for_delivery', na=False).sum())

        results.append({
            'Week': sheet,
            'Total': total,
            'Delivered': delivered,
            'Awaiting Pickup': awaiting_pickup,
            'Out for Delivery': out_for_delivery,
        })

    # ── Terminal output ──────────────────────────────────────────────────────
    print(f"File: {os.path.basename(filepath)}\n")
    print(f"{'WEEK':<8} {'TOTAL':>8}   {'DELIVERED':>12}   {'AWAITING PICKUP':>16}   {'OUT FOR DELIVERY':>17}")
    print("-" * 75)
    for r in results:
        t = r['Total']
        print(
            f"{r['Week']:<8} {t:>8}   "
            f"{r['Delivered']:>6} ({pct(r['Delivered'], t):>5.1f}%)   "
            f"{r['Awaiting Pickup']:>6} ({pct(r['Awaiting Pickup'], t):>5.1f}%)   "
            f"{r['Out for Delivery']:>6} ({pct(r['Out for Delivery'], t):>5.1f}%)"
        )
    print()

    # ── Write Cohorts sheet ──────────────────────────────────────────────────
    wb = load_workbook(filepath)

    if 'Cohorts' in wb.sheetnames:
        del wb['Cohorts']

    ws = wb.create_sheet('Cohorts')

    headers = ['Week', 'Total Orders', 'Delivered', 'Del %', 'Awaiting Pickup', 'AP %', 'Out for Delivery', 'OFD %']
    header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, r in enumerate(results, 2):
        t = r['Total']
        row = [
            r['Week'], t,
            r['Delivered'],        pct(r['Delivered'], t),
            r['Awaiting Pickup'],  pct(r['Awaiting Pickup'], t),
            r['Out for Delivery'], pct(r['Out for Delivery'], t),
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='center')

        # shade the % cells light blue
        for col_idx in [4, 6, 8]:
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type='solid', fgColor='DCE6F1')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 6

    wb.save(filepath)
    print(f"Cohorts tab saved.")


if __name__ == '__main__':
    filepath = get_latest_xlsx()

    if not filepath:
        print("No .xlsx file found in this folder. Drop your Excel file here and try again.")
        input("\nPress Enter to close...")
        sys.exit(1)

    run(filepath)
    input("Done! Press Enter to close...")
